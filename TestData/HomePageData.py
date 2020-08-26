import openpyxl
class HomePageData:

    test_HomePage_Data = [{"firstname":"graham dale", "email":"graham_dale@yahoo.com", "gender":"Male"},{"firstname":"Balto Dale", "email":"irish0341@gmail.com", "gender":"Female"}]



    @staticmethod
    def getTestData(test_case_name):
        Dict = {}
        book = openpyxl.load_workbook("C:\\Users\\Home Laptop\\PycharmProjects\\pythonSelFramework\\TestData\\ExcelDemo.xlsX")
        sheet = book.active
        for i in range(1, sheet.max_row + 1):  # to get rows
            if sheet.cell(row=i, column=1).value == test_case_name:

                for j in range(2, sheet.max_column + 1):  # to get columns

                    Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
        return[Dict]

